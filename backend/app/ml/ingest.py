"""BhoomiRashi workbook ingestion pipeline.

Two layers:
- ``parse_workbook``: pure, side-effect-free parsing/cleaning of an XLSX file
  into validated parcel/party records plus a data-quality report. Unit-testable
  without a database.
- ``run_ingestion``: idempotent persistence into the ``imported_*`` staging
  tables (raw values preserved, normalized fields derived). Rerunning with the
  same source file replaces that file's rows instead of duplicating them.

Usage (offline / administrative only — never at application startup):
    python -m app.ml.ingest /path/to/workbook.xlsx
"""

import argparse
import asyncio
import logging
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import Optional

from sqlalchemy import delete

from app.ml.normalize import (
    cell_text,
    is_compound_survey,
    map_land_type,
    map_ownership_status,
    normalize_party_type,
    normalize_survey_number,
    normalize_text,
    parcel_dedupe_key,
    party_dedupe_key,
    parse_area_hectares,
    raw_text,
    strip_quotes,
    survey_number_head,
)

logger = logging.getLogger(__name__)


@dataclass
class ParsedParcel:
    source_sno: str
    dedupe_key: str
    raw: dict
    district_norm: Optional[str]
    sub_district_norm: Optional[str]
    village_norm: Optional[str]
    survey_number_norm: Optional[str]
    survey_head: Optional[int]
    compound: bool
    area_hectares: Optional[Decimal]
    land_type_mapped: str
    land_type_mapped_ok: bool
    ownership_mapped: str
    ownership_mapped_ok: bool
    land_category: Optional[str]
    land_nature_label: Optional[str]


@dataclass
class ParsedParty:
    source_sno: str
    dedupe_key: str
    raw: dict
    name_norm: Optional[str]
    address_norm: Optional[str]
    party_type: str
    party_type_mapped_ok: bool
    area_hectares: Optional[Decimal]


@dataclass
class IngestionReport:
    source_file: str
    land_rows_seen: int = 0
    land_rows_loaded: int = 0
    land_rows_rejected: int = 0
    land_rows_duplicate: int = 0
    party_rows_seen: int = 0
    party_rows_loaded: int = 0
    party_rows_rejected: int = 0
    party_rows_duplicate: int = 0
    party_rows_unlinked: int = 0
    document_title: Optional[str] = None
    document_publish_date: Optional[str] = None
    rejected_land: list = field(default_factory=list)
    rejected_party: list = field(default_factory=list)


def parse_workbook(
    path: str | Path,
) -> tuple[list[ParsedParcel], list[ParsedParty], IngestionReport]:
    """Parse and validate the workbook. Pure function (no DB, no writes)."""
    import openpyxl

    report = IngestionReport(source_file=Path(path).name)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    # --- Document Information ---
    if "Document Information" in wb.sheetnames:
        ws = wb["Document Information"]
        info = {cell_text(r[0]): cell_text(r[1]) for r in ws.iter_rows(values_only=True)}
        report.document_title = info.get("Title")
        report.document_publish_date = info.get("Tentative Publish Date")

    # --- Land Details ---
    parcels: list[ParsedParcel] = []
    seen_keys: set[str] = set()
    ws = wb["Land Details"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        if all(c is None for c in row):
            continue
        report.land_rows_seen += 1
        sno = cell_text(row[0])
        if not sno:
            report.land_rows_rejected += 1
            report.rejected_land.append({"row": report.land_rows_seen, "reason": "missing S.No"})
            continue
        raw = {
            "district": raw_text(row[1]),
            "sub_district": raw_text(row[2]),
            "village": raw_text(row[3]),
            "survey_number": raw_text(row[4]),
            "area": raw_text(row[5]),
            "description": raw_text(row[6]),
            "land_type": raw_text(row[7]),
            "land_nature": raw_text(row[8]),
            "land_category": raw_text(row[9]),
            "additional_details": raw_text(row[10]) if len(row) > 10 else None,
        }
        district = cell_text(row[1])
        sub_district = cell_text(row[2])
        village = cell_text(row[3])
        survey_number = cell_text(row[4])

        area_value, area_error = parse_area_hectares(raw_text(row[5]))
        if area_error:
            report.land_rows_rejected += 1
            report.rejected_land.append({"sno": sno, "reason": area_error})
            continue
        if not (district and sub_district and village and survey_number):
            report.land_rows_rejected += 1
            report.rejected_land.append(
                {"sno": sno, "reason": "missing district/sub-district/village/survey number"}
            )
            continue

        key = parcel_dedupe_key(district, sub_district, village, survey_number)
        if key in seen_keys:
            report.land_rows_duplicate += 1
            continue
        seen_keys.add(key)

        land_type, land_type_ok = map_land_type(strip_quotes(cell_text(row[7])))
        ownership, ownership_ok = map_ownership_status(strip_quotes(cell_text(row[8])))
        parcels.append(
            ParsedParcel(
                source_sno=sno,
                dedupe_key=key,
                raw=raw,
                district_norm=normalize_text(district),
                sub_district_norm=normalize_text(sub_district),
                village_norm=normalize_text(village),
                survey_number_norm=normalize_survey_number(survey_number),
                survey_head=survey_number_head(survey_number),
                compound=is_compound_survey(survey_number),
                area_hectares=area_value,
                land_type_mapped=land_type,
                land_type_mapped_ok=land_type_ok,
                ownership_mapped=ownership,
                ownership_mapped_ok=ownership_ok,
                land_category=normalize_text(raw_text(row[9])),
                land_nature_label=normalize_text(strip_quotes(cell_text(row[8]))),
            )
        )
        report.land_rows_loaded += 1

    # --- Land Parties ---
    parties: list[ParsedParty] = []
    seen_party_keys: set[str] = set()
    valid_snos = {p.source_sno for p in parcels}
    ws = wb["Land Parties"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        if all(c is None for c in row):
            continue
        report.party_rows_seen += 1
        sno = cell_text(row[0])
        name = cell_text(row[2])
        address = cell_text(row[3])
        if not sno or not name:
            report.party_rows_rejected += 1
            report.rejected_party.append(
                {"row": report.party_rows_seen, "reason": "missing Source S.No or Name"}
            )
            continue
        area_value, area_error = parse_area_hectares(raw_text(row[5]) if len(row) > 5 else None)
        if area_error and area_error != "missing area":
            report.party_rows_rejected += 1
            report.rejected_party.append({"sno": sno, "name": name[:40], "reason": area_error})
            continue
        key = party_dedupe_key(sno, name, address)
        if key in seen_party_keys:
            report.party_rows_duplicate += 1
            continue
        seen_party_keys.add(key)
        if sno not in valid_snos:
            report.party_rows_unlinked += 1
        party_type, party_type_ok = normalize_party_type(
            cell_text(row[4]) if len(row) > 4 else None
        )
        parties.append(
            ParsedParty(
                source_sno=sno,
                dedupe_key=key,
                raw={
                    "name": raw_text(row[2]),
                    "address": raw_text(row[3]),
                    "type": raw_text(row[4]) if len(row) > 4 else None,
                    "area": raw_text(row[5]) if len(row) > 5 else None,
                },
                name_norm=normalize_text(name),
                address_norm=normalize_text(address),
                party_type=party_type,
                party_type_mapped_ok=party_type_ok,
                area_hectares=area_value,
            )
        )
        report.party_rows_loaded += 1

    wb.close()
    return parcels, parties, report


async def run_ingestion(db, path: str | Path) -> IngestionReport:
    """Idempotently load the workbook into staging tables.

    Rows from a previous run of the *same source file* are deleted first, so
    reruns never duplicate. Rows from other source files are untouched.
    """
    from app.models.import_staging import ImportedLandDetail, ImportedLandParty

    parcels, parties, report = parse_workbook(path)

    source_file = report.source_file
    await db.execute(delete(ImportedLandParty).where(ImportedLandParty.source_file == source_file))
    await db.execute(
        delete(ImportedLandDetail).where(ImportedLandDetail.source_file == source_file)
    )

    detail_ids: dict[str, object] = {}
    party_counts: dict[str, int] = {}
    for party in parties:
        party_counts[party.source_sno] = party_counts.get(party.source_sno, 0) + 1

    for parcel in parcels:
        record = ImportedLandDetail(
            source_file=source_file,
            source_sno=parcel.source_sno,
            dedupe_key=parcel.dedupe_key,
            raw_district=parcel.raw["district"],
            raw_sub_district=parcel.raw["sub_district"],
            raw_village=parcel.raw["village"],
            raw_survey_number=parcel.raw["survey_number"],
            raw_area=parcel.raw["area"],
            raw_description=parcel.raw["description"],
            raw_land_type=parcel.raw["land_type"],
            raw_land_nature=parcel.raw["land_nature"],
            raw_land_category=parcel.raw["land_category"],
            raw_additional_details=parcel.raw["additional_details"],
            district_norm=parcel.district_norm,
            sub_district_norm=parcel.sub_district_norm,
            village_norm=parcel.village_norm,
            survey_number_norm=parcel.survey_number_norm,
            survey_number_head=parcel.survey_head,
            is_compound_survey=parcel.compound,
            area_hectares=parcel.area_hectares,
            land_type_mapped=parcel.land_type_mapped,
            land_type_mapped_ok=parcel.land_type_mapped_ok,
            ownership_status_mapped=parcel.ownership_mapped,
            ownership_status_mapped_ok=parcel.ownership_mapped_ok,
            land_category=parcel.land_category,
            land_nature_label=parcel.land_nature_label,
            party_count=party_counts.get(parcel.source_sno, 0),
        )
        db.add(record)
        await db.flush()
        detail_ids[parcel.source_sno] = record.id

    for party in parties:
        db.add(
            ImportedLandParty(
                source_file=source_file,
                source_sno=party.source_sno,
                dedupe_key=party.dedupe_key,
                raw_name=party.raw["name"],
                raw_address=party.raw["address"],
                raw_type=party.raw["type"],
                raw_area=party.raw["area"],
                name_norm=party.name_norm,
                address_norm=party.address_norm,
                party_type=party.party_type,
                party_type_mapped_ok=party.party_type_mapped_ok,
                area_hectares=party.area_hectares,
                imported_detail_id=detail_ids.get(party.source_sno),
            )
        )

    await db.commit()
    logger.info(
        "Ingestion of %s complete: %d parcels, %d parties (%d rejected land rows, %d rejected party rows)",
        source_file,
        report.land_rows_loaded,
        report.party_rows_loaded,
        report.land_rows_rejected,
        report.party_rows_rejected,
    )
    return report


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Ingest a BhoomiRashi workbook into NLAMS staging tables"
    )
    parser.add_argument("workbook", help="Path to the XLSX workbook")
    parser.add_argument(
        "--db-url", default=None, help="Async database URL (defaults to app settings)"
    )
    args = parser.parse_args()

    import json

    from app.db.session import get_engine

    async def _run():
        from sqlalchemy.ext.asyncio import async_sessionmaker, AsyncSession, create_async_engine
        from app.core.config import settings

        eng = create_async_engine(args.db_url or settings.DATABASE_URL)
        factory = async_sessionmaker(eng, class_=AsyncSession, expire_on_commit=False)
        async with factory() as session:
            report = await run_ingestion(session, args.workbook)
        await eng.dispose()
        print(json.dumps(report.__dict__, indent=2, default=str))

    asyncio.run(_run())


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    main()
